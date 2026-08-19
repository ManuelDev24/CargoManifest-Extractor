from typing import List
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from core.models import ManifestRecord
from validators.manifest_validator import ValidationReport


class ExcelExporter:
    @staticmethod
    def export(records: List[ManifestRecord], validation: ValidationReport, output_path: str):
        out = Path(output_path)
        out.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        ws = wb.active
        ws.title = "Manifiesto Normalizado"

        headers = list(records[0].to_dict().keys())
        ws.append(headers)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        center = Alignment(horizontal="center")

        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

        for rec in records:
            ws.append(list(rec.to_dict().values()))

        # auto width
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 60)

        # validation sheet
        ws2 = wb.create_sheet(title="Validacion")
        ws2.append(["Total", "Validos", "Errores", "Advertencias"]) 
        ws2.append([validation.total_records, validation.valid_records, validation.error_count, validation.warning_count])

        ws2.append([])
        ws2.append(["ID", "B/L", "Equipo", "Campo", "Severidad", "Mensaje"])
        for issue in validation.issues:
            ws2.append([issue.record_index, issue.bl_number, issue.equipment_id, issue.field_name, issue.severity, issue.message])

        wb.save(out)
